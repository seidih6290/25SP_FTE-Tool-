# -*- coding: utf-8 -*-
"""
Takes data from deanDailyCsar and FTE_Tier to determine classes FTE.
Allows the user to Get Course Enrollment, and get FTE by Division, Instructor,
and Course.

GROUP A
Thuan Chau, Karen Brown, Harley Coughlin,Teresa Hearn, Shiane Ransford,
Latoya Winston

04/28/2025

CSC-221-001

M7GroupAnBPro

"""

import traceback

# import re
# import os
import pandas as pd

# import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill

# import openpyxl
from openpyxl.utils import get_column_letter


def menu():
    """
    Displays the menu options.

    Returns
    -------
    None.F

    """
    print()
    print("=" * 20 + "Menu" + "=" * 20)
    print('1) Enter "Sec Divisions" code ')
    print("2) Get course Enrollment Percentage")
    print("3) FTE by Division")
    print("4) FTE per instructor (for specific Div)")
    print("5) FTE per course (for specific Div)")
    print("6) Exit")
    print("=" * 44)


def readfile():
    """
    Generates the dataframe and then sorts it.

    Returns
    -------
    groups : dataframe
        the sorted dataframe of the file.

    """
    try:
        # reads the deansDailyCsar.csv and unique_deansDailyCsar_FTE files in
        # to a dataframe
        file_in = pd.read_csv("deanDailyCsar.csv")
        fte_file_in = pd.read_excel("unique_deansDailyCsar_FTE.xlsx")

        # merge prior dataframes
        # Extract Course Code from Sec Name if not already done
        if "Course Code" not in file_in.columns:
            file_in["Course Code"] = file_in["Sec Name"].str.extract(
                r"([A-Z]{3}-\d{3})"
            )

        # Also create Course Code in credits_df
        if "Course Code" not in fte_file_in.columns:
            fte_file_in["Course Code"] = fte_file_in["Sec Name"].str.extract(
                r"([A-Z]{3}-\d{3})"
            )

        # Merge only needed columns from credits_df
        merged_df = pd.merge(
            file_in,
            fte_file_in[["Course Code", "Contact Hours"]],
            how="left",
            on="Course Code",
        )

        merged_df["Contact Hours"] = pd.to_numeric(
            merged_df["Contact Hours"], errors="coerce"
        )
        merged_df["FTE Count"] = pd.to_numeric(merged_df["FTE Count"],
                                               errors="coerce")

        # Calculate Total FTE
        merged_df["Total FTE"] = (
            (merged_df["Contact Hours"] * 16 * merged_df["FTE Count"]) / 512
        ).round(3)

        # sorts the dataframe by sec divisions, sec name
        # and sec faculty info and assigns it to groups
        groups = merged_df.sort_values(
            ["Sec Divisions", "Sec Name", "Sec Faculty Info"]
        )
        #print(groups.head(10))
        return groups

    except FileNotFoundError:
        groups = []
        print("File Missing!")
        return groups


def auto_format_excel(filename):
    """
    Auto-formats the column widths of an Excel file to fit the content.

    Parameters
    ----------
    filename : str
        The name of the Excel file to format.

    Returns
    -------
    None.
    """
    try:
        # Load the workbook and select the active sheet
        wb = load_workbook(filename)
        ws = wb.active

        assert isinstance(ws, Worksheet)
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            assert column[0].column is not None
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                max_length = max(len(str(cell.value)), max_length)

            # Adjust width with a little extra space
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the formatted workbook
        wb.save(filename)
        print(f"Formatted {filename} with auto-adjusted column widths")

    except Exception as err:
        print(f"Error formatting Excel file: {str(err)}")


def sec_divisions(file_in):
    """
    Allows user to enter sec divisions to search for.

    Parameters
    ----------
    file_in : dataframe
        Contains the information for each division.

    Returns
    -------
    None.

    """
    try:
        # access and displays the available sec divisons
        # to choose from in rows of 4.
        print()
        print("Available Sec Divisions: \n")
        sec_group = sorted(file_in["Sec Divisions"].dropna().unique())

        # Display divisions in rows of 4
        for i in range(0, len(sec_group), 4):
            row_items = sec_group[i: i + 4]
            for x in row_items:
                print(f"-{x}", end=" ")
            print()

        # Get and process users input
        sec_input = input("\nEnter Sec Divisions separated by commas or ALL: ")
        sec_input = sec_input.upper().strip()

        if sec_input == "ALL":  # Check for ALL before splitting
            divisions_to_process = sec_group
        else:
            # Split only if not ALL
            divisions_to_process = sec_input.split(",")
            divisions_to_process = [div.strip() for div
                                    in divisions_to_process]

        # Validate and process divisions
        for div in divisions_to_process:
            if div.upper() not in [d.upper() for d in sec_group]:
                print(f"\nWarning: Division '{div}' not found")

        for division in divisions_to_process:
            if division.upper() in [d.upper() for d in sec_group]:
                # Extract rows for selected division into a new dataframe
                # Convert division name to lowercase for dataframe name
                df_names = division.lower()

                # Filter for the division and remove "Course Code" column if
                # it exists, ensure "Contact Hours" is included
                if "Course Code" in file_in.columns:
                    # Get all columns except "Course Code"
                    columns_to_keep = [
                        col for col in file_in.columns if col != "Course Code"
                    ]
                    df_name = file_in[file_in["Sec Divisions"] == division][
                        columns_to_keep
                    ].copy()
                else:
                    df_name = file_in[file_in["Sec Divisions"] == division
                                      ].copy()

                # Check if "Contact Hours" is in the dataframe
                if "Contact Hours" not in df_name.columns:
                    print("\nWarning: 'Contact Hours' column not found in the "
                          "data")

                # Create Excel filename (lowercase)
                excel_filename = f"{division.lower()}.xlsx"

                # Write to Excel
                df_name.to_excel(excel_filename, index=False)

                # Auto-format the Excel file columns
                auto_format_excel(excel_filename)

                print(f"\nCreated DataFrame '{df_names}' with {len(df_name)}"
                      " rows")
                print(f"Saved to file: {excel_filename}")

    except TypeError:
        print("Missing information from file. Check to be sure the file is not"
              " missing.")
    except Exception as err:
        print("Error: " + str(err))


def option2_enrollment(df):
    """
    Parameters
    ----------
    df : dataframe
        file data fram deansDailyCsar.csv

    Returns
    -------
    Returns course enrollment percentage

    """
    course_code = True

    # Avoid potential unbound issues
    filtered_df = pd.DataFrame()
    course_input = None
    while course_code:
        course_input = input(
            "Enter course code (e.g., ACA-120) or type 'back' to return: "
        ).strip()

        if course_input.lower() == "back":
            return

        # Filter rows in 'Sec Name' containing the course code
        # (case insensitive)
        filtered_df = df[
            df["Sec Name"].str.contains(course_input, case=False, na=False)
        ]

        if filtered_df.empty:
            print(
                "Course not found. Please re-enter the course code or"
                "type 'back' to return to the main menu."
            )
        else:
            course_code = False

    # For face-to-face sections, duplicate rows may exist.
    # Here, we drop duplicate rows based on 'Sec Name'.
    # This assumes online sections (which contain a '9' in the section number)
    # are unique or do not duplicate.
    filtered_df = filtered_df.drop_duplicates(subset="Sec Name")

    # Define a function to calculate enrollment percentage for a row
    def calc_enrollment(row):
        try:
            cap = float(row["Capacity"])
            fte = float(row["FTE Count"])

            if cap == 0:
                return "0%"

            percentage = (fte / cap) * 100
            return f"{percentage:.2f}%"

        except (ValueError, TypeError, ZeroDivisionError):
            return "N/A%"

    # Calculate and add the Enrollment Percentage column
    filtered_df["Enrollment Percentage"] = filtered_df.apply(calc_enrollment,
                                                             axis=1)

    # Create the output DataFrame with the required columns
    output_columns = [
        "Sec Name",
        "X Sec Delivery Method",
        "Meeting Times",
        "Capacity",
        "FTE Count",
        "Total FTE",
        "Sec Faculty Info",
        "Enrollment Percentage",
    ]
    output_df = filtered_df[output_columns]

    # Determine the file name based on the course code entered
    assert course_input is not None
    file_code = course_input.replace("-", "").lower()
    file_name = f"{file_code}_per.xlsx"

    # Put output.df into an excel file
    output_df.to_excel(file_name, index=False)

    # Adjust column widths with openpyxl workbook
    wb = load_workbook(file_name)
    ws = wb.active
    assert isinstance(ws, Worksheet)

    # Adjust column widths dynamically
    for column_cells in ws.columns:
        assert column_cells[0].column is not None
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        column_name = column_cells[0].value  # Header value

        for cell in column_cells:
            if cell.value:
                max_length = max(len(str(cell.value)), max_length)

        # If column is "Meeting Times", make it extra wide
        if column_name == "Meeting Times":
            ws.column_dimensions[column_letter].width = max_length + 1
        else:
            ws.column_dimensions[column_letter].width = 25
    # Save the File
    wb.save(file_name)
    print(f"Created '{file_name}' with enrollment data.")


def division_fte(file_in):
    """
    Analyze FTE by Division and export to a sheet in a division-specific report
    file.

    Parameters
    ----------
    file_in : pandas.DataFrame
        Input DataFrame containing course information.

    Returns
    -------
    None
    """

    print()
    # Get unique division codes
    divisions = sorted(file_in["Sec Divisions"].dropna().unique())

    # Display available divisions
    print("Available Division Codes:")
    for i in range(0, len(divisions), 4):
        row = divisions[i: i + 4]
        print("  ".join(f"{div}" for div in row))

    # Get division code from user
    div_code = input("\nEnter Division Code: ").strip()

    if not div_code:
        print("Please enter a valid division code.")
        return

    # Convert to uppercase for case-insensitive comparison
    div_code = div_code.upper()

    # Check if division exists (case-insensitive comparison)
    valid_divisions = [div.upper() for div in divisions]
    if div_code not in valid_divisions:
        print(f"Division '{div_code}' not found. Please check the code and try"
              " again.")
        return

    # Read FTE tier data
    fte_data = pd.read_excel("FTE_Tier.xlsx")

    # Create a lookup dictionary for faster access
    fte_lookup = {
        row["Prefix/Course ID"]: row["New Sector"]
        for _, row in fte_data.iterrows()
        if not pd.isna(row.loc["Prefix/Course ID"])
    }

    # Get the actual division code with correct case
    actual_div = divisions[valid_divisions.index(div_code)]

    try:
        # Filter data for the selected division
        div_data = file_in[file_in["Sec Divisions"] == actual_div].copy()

        if len(div_data) == 0:
            print("No data found for this division.")
            return

        # Add course code column
        div_data["Course Code"] = div_data["Sec Name"].str.extract(
            r"([A-Z]+-\d+)")

        # Sort by Course Code and Sec Name
        div_data = div_data.sort_values(["Course Code", "Sec Name"])

        # Base value for FTE calculation
        base_fte_value = 1926

        # Create output list to store rows
        output_rows = []
        current_course = None
        course_total_fte = 0
        first_row = True

        # Track grand totals
        grand_total_fte = 0
        grand_total_original_fte = 0

        # Process each row
        for _, row in div_data.iterrows():
            course = row["Course Code"]

            # If new course and not first course, add total for previous course
            if course != current_course and current_course is not None:
                output_rows.append(
                    {
                        "Division": "",
                        "Course Code": "Total",
                        "Sec Name": "",
                        "X Sec Delivery Method": "",
                        "Meeting Times": "",
                        "Capacity": "",
                        "FTE Count": "",
                        "Sec Faculty Info": "",
                        "Total FTE": "",
                        "Enrollment Per": "",
                        "Generated FTE": course_total_fte,
                    }
                )

                # Add to grand total
                grand_total_fte += course_total_fte

                course_total_fte = 0

            # Get section prefix (first 3 characters of section name)
            sec = row["Sec Name"][:3] if not pd.isna(row["Sec Name"]) else ""

            # Look up new sector value from the dictionary
            new_sector_value = fte_lookup.get(sec, 0)

            # Calculate adjusted FTE for the current row
            total_fte = float(row["Total FTE"]) if pd.notna(row["Total FTE"]) else 0

            # Add to grand total of original FTE
            grand_total_original_fte += total_fte

            # Calculate adjusted_fte
            adjusted_fte = total_fte * (new_sector_value + base_fte_value)

            # Calculate enrollment percentage
            enrollment_per = ""
            if (
                pd.notna(row["Capacity"])
                and pd.notna(row["FTE Count"])
                and float(row["Capacity"]) > 0
            ):
                enrollment_per = (
                    float(row["FTE Count"]) / float(row["Capacity"])
                ) * 100
                enrollment_per = "{:.2f}%".format(round(enrollment_per, 2))

            # Add current row with enrollment percentage and generated FTE
            output_rows.append(
                {
                    "Division": actual_div if first_row else "",
                    "Course Code": course if course != current_course else "",
                    "Sec Name": row["Sec Name"],
                    "X Sec Delivery Method": row["X Sec Delivery Method"],
                    "Meeting Times": row["Meeting Times"],
                    "Capacity": row["Capacity"],
                    "FTE Count": row["FTE Count"],
                    "Sec Faculty Info": row["Sec Faculty Info"],
                    "Total FTE": row["Total FTE"],
                    "Enrollment Per": enrollment_per,
                    "Generated FTE": adjusted_fte,
                }
            )

            # Add to course total generated fte
            course_total_fte += adjusted_fte

            current_course = course
            first_row = False

        # Add total for last course
        if current_course is not None:
            output_rows.append(
                {
                    "Division": "",
                    "Course Code": "Total",
                    "Sec Name": "",
                    "X Sec Delivery Method": "",
                    "Meeting Times": "",
                    "Capacity": "",
                    "FTE Count": "",
                    "Sec Faculty Info": "",
                    "Total FTE": "",
                    "Enrollment Per": "",
                    "Generated FTE": course_total_fte,
                }
            )

            # Add last course total to grand total
            grand_total_fte += course_total_fte

        # Add grand total row with formatted total
        output_rows.append(
            {
                "Division": "",
                "Course Code": "DIVISION TOTAL",
                "Sec Name": "",
                "X Sec Delivery Method": "",
                "Meeting Times": "",
                "Capacity": "",
                "FTE Count": "",
                "Sec Faculty Info": "",
                # Ensure it's a float
                "Total FTE": float(grand_total_original_fte),
                "Enrollment Per": "",
                # Ensure it's a float
                "Generated FTE": float(grand_total_fte),
            }
        )

        # Format Generated FTE column with '$' before the number
        # and a comma every 3 digits
        for row in output_rows:
            # Format Generated FTE if it's a number
            if isinstance(row["Generated FTE"], (int, float)):
                row["Generated FTE"] = "${:,.2f}".format(row["Generated FTE"])

            # Format Total FTE if it's a number (including the grand total)
            if isinstance(row["Total FTE"], (int, float)):
                row["Total FTE"] = "{:.2f}".format(row["Total FTE"])

        # Convert to DataFrame
        output_df = pd.DataFrame(output_rows)

        # Create Excel file
        excel_filename = f"{actual_div.lower()}_fte.xlsx"

        # Write to Excel
        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
            output_df.to_excel(writer, sheet_name="Division Analysis",
                               index=False)

            # Format the worksheet
            worksheet = writer.sheets["Division Analysis"]

            # Adjust column widths
            for idx, col in enumerate(output_df.columns):
                max_length = (
                    max(output_df[col].astype(str).apply(len).max(),
                        len(str(col))) + 2
                )
                worksheet.column_dimensions[chr(65 + idx)].width = max_length

            # Apply bold formatting to the Division Total row
            last_row = (
                len(output_df) + 1
            )  # +1 because Excel rows are 1-indexed and we have a header row
            for col in range(1, len(output_df.columns) + 1):
                cell = worksheet.cell(row=last_row, column=col)
                cell.font = Font(bold=True)

            # Add a bottom border to cells before the grand total
            second_last_row = last_row - 1
            # Make sure there are rows before the total
            if second_last_row > 1:
                for col in range(1, len(output_df.columns) + 1):
                    cell = worksheet.cell(row=second_last_row, column=col)
                    cell.border = Border(bottom=Side(style="thin"))

            # Add background color to the grand total row
            for col in range(1, len(output_df.columns) + 1):
                cell = worksheet.cell(row=last_row, column=col)
                cell.fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )

        print("\nAnalysis for division: {}".format(actual_div))
        print("Results exported to {}".format(excel_filename))
        print("Division Total Original FTE: {:.2f}".format(grand_total_original_fte))
        print("Division Total Generated FTE: ${:,.2f}".format(grand_total_fte))

    except Exception:
        print("Error processing data")
        print(traceback.format_exc())
        return

    return


def clean_name_for_search(name):
    """
    Standardize name format for searching.
    Removes periods and extra spaces.

    Parameters
    ----------
    name : str
        Name to clean

    Returns
    -------
    str
        Cleaned name for comparison
    """
    return name.replace(".", "").strip().lower()


def clean_instructor_name(name):
    """
    Clean instructor name for file naming.
    Handles different formats like "H Seidi", "H. Seidi", etc.

    Parameters
    ----------
    name : str
        Instructor name

    Returns
    -------
    str
        Cleaned name formatted for filename
    """
    # Split by comma first if it exists
    if "," in name:
        last_name, first_part = name.split(",", 1)
        # Clean up the last name and first initial
        last_name = last_name.strip().lower()
        # Get first character and remove any periods
        first_initial = first_part.strip().replace(".", "")[0].lower()
    else:
        # Handle space-separated names
        parts = name.split()
        last_name = parts[-1].lower()
        # Get first character and remove any periods
        first_initial = parts[0].replace(".", "")[0].lower()

    return f"{last_name}{first_initial}_FTE.xlsx"


def clean_course_code(code):
    """
    Clean course code for file naming.
    Removes dashes and standardizes format.
    """
    # Remove dash and convert to lowercase
    clean_code = code.replace("-", "").lower()
    return f"{clean_code}_FTE.xlsx"


def fte_per_course(file_in):
    """
    Calculate and export FTE data for a specific course.
    """

    print()
    # Step 1: Extract course codes from section names using regex
    file_in["Course Code"] = file_in["Sec Name"].str.extract(r"([A-Z]+-\d+)")
    course_codes = sorted(file_in["Course Code"].dropna().unique())

    while True:
        # Step 2a: Prompt user for course code input
        print("\nEnter course code (e.g., CSC-121) or type 'back' to return to"
              " main menu:")
        course_input = input("Course code: ").strip().upper()

        # Step 2b: Allow user to exit function
        if course_input.lower() == "back":
            return

        # Step 2c: Find matching courses based on partial or complete input
        matching_courses = [c for c in course_codes if course_input in c]

        # Step 2d: Handle case when no matching courses found
        if not matching_courses:
            print(f"No course found with code '{course_input}'.")
            continue

        # Step 2e: Select specific course (or have user select if multiple
        # matches)
        if len(matching_courses) == 1:
            selected_course = matching_courses[0]
        else:
            print("\nMultiple courses found:")
            for i, course in enumerate(matching_courses, 1):
                print(f"{i}. {course}")
            selection = input("\nEnter number to select course: ")
            if (
                not selection.isdigit()
                or int(selection) < 1
                or int(selection) > len(matching_courses)
            ):
                print("Invalid selection.")
                continue
            selected_course = matching_courses[int(selection) - 1]

        print(f"\nProcessing data for course: {selected_course}")

        try:
            # Step 3a: Filter data for selected course
            course_data = file_in[file_in["Course Code"] == selected_course].copy()
            if course_data.empty:
                print("No sections found for this course.")
                continue

            # Step 3b: Remove duplicate sections
            course_data = course_data.drop_duplicates(subset="Sec Name")

            # Step 3c: Load FTE tier data for calculations
            fte_data = pd.read_excel("FTE_Tier.xlsx")
            fte_lookup = {
                row["Prefix/Course ID"]: row["New Sector"]
                for _, row in fte_data.iterrows()
                if pd.notna(row.loc["Prefix/Course ID"])
            }

            # Step 3d: Sort data by section name
            course_data = course_data.sort_values("Sec Name")
            base_fte_value = 1926

            # Step 4a: Initialize variables for calculating FTE
            output_rows = []
            total_generated_fte = 0
            total_original_fte = 0  # Track sum of original Total FTE

            # Step 4b: Process each section to calculate FTE metrics
            for _, row in course_data.iterrows():
                # Extract section prefix (first 3 chars of section name)
                sec_prefix = row["Sec Name"][:3] if not pd.isna(row["Sec Name"]) else ""

                # Look up new sector value for this prefix
                new_sector_value = fte_lookup.get(sec_prefix, 0)

                # Calculate total FTE for this section
                total_fte = float(row["Total FTE"]) if pd.notna(row["Total FTE"]) else 0

                # Add to total original FTE
                total_original_fte += total_fte

                # Calculate generated FTE using formula
                generated_fte = total_fte * (new_sector_value + base_fte_value)

                # Calculate enrollment percentage
                enrollment_per = ""
                if (
                    pd.notna(row["Capacity"])
                    and pd.notna(row["FTE Count"])
                    and float(row["Capacity"]) > 0
                ):
                    enrollment_per = round(
                        (float(row["FTE Count"]) / float(row["Capacity"])) * 100, 2
                    )

                # Add section data to output rows
                output_rows.append(
                    {
                        "Course Code": selected_course if len(output_rows) == 0 else "",
                        "Sec Name": row["Sec Name"],
                        "X Sec Delivery Method": row["X Sec Delivery Method"],
                        "Sec Faculty Info": row["Sec Faculty Info"],
                        "Meeting Times": row["Meeting Times"],
                        "Capacity": row["Capacity"],
                        "FTE Count": row["FTE Count"],
                        "Total FTE": total_fte,
                        "Enrollment Per": "{:.2f}%".format(enrollment_per)
                        if enrollment_per != ""
                        else "",
                        "Generated FTE": generated_fte,
                    }
                )

                # Add current section's FTE to course total
                total_generated_fte += generated_fte

            # Step 4c: Add a summary row with course totals
            output_rows.append(
                {
                    "Course Code": "COURSE TOTAL",
                    "Sec Name": "",
                    "X Sec Delivery Method": "",
                    "Sec Faculty Info": "",
                    "Meeting Times": "",
                    "Capacity": "",
                    "FTE Count": "",
                    "Total FTE": total_original_fte,  # Add total original FTE
                    "Enrollment Per": "",
                    "Generated FTE": total_generated_fte,
                }
            )

            # Format the numeric values
            for row in output_rows:
                # Format Generated FTE if it's a number
                if isinstance(row["Generated FTE"], (int, float)):
                    row["Generated FTE"] = "${:,.2f}".format(row["Generated FTE"])

                # Format Total FTE if it's a number (including the total row)
                if isinstance(row["Total FTE"], (int, float)):
                    row["Total FTE"] = "{:.2f}".format(row["Total FTE"])

            # Step 5a: Convert to DataFrame for export
            output_df = pd.DataFrame(output_rows)

            # Step 5b: Create Excel file with course code as name
            file_name = f"{selected_course.replace('-', '').lower()}_FTE.xlsx"

            # Create a writer for Excel output
            with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
                output_df.to_excel(writer, index=False,
                                   sheet_name="Course Analysis")

                # Get the worksheet and apply formatting
                worksheet = writer.sheets["Course Analysis"]

                # Adjust column widths
                for idx, col in enumerate(output_df.columns):
                    max_length = (
                        max(output_df[col].astype(str).apply(len).max(),
                            len(str(col))) + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = max_length

                # Apply formatting to the Course Total row (bold and
                # background color)
                last_row = len(output_df) + 1
                for col in range(1, len(output_df.columns) + 1):
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E0E0E0", end_color="E0E0E0",
                        fill_type="solid"
                    )

                # Add a bottom border to cells before the total row
                second_last_row = last_row - 1
                # Make sure there are rows before the total
                if second_last_row > 1:
                    for col in range(1, len(output_df.columns) + 1):
                        cell = worksheet.cell(row=second_last_row, column=col)
                        cell.border = Border(bottom=Side(style="thin"))

            # Display summary of results
            print(f"\nAnalysis for course: {selected_course}")
            print(f"Found {len(course_data)} sections")
            print(f"Results exported to {file_name}")
            print("Course Total Original FTE: {:.2f}".format(total_original_fte))
            print("Course Total Generated FTE: ${:,.2f}".format(total_generated_fte))
            break

        except Exception as e:
            print("Error processing course data")
            print(str(e))
            print(traceback.format_exc())
            continue
