"""
web_functions.py

This module provides backend processing for the Faculty FTE Report Generator
Streamlit application. It contains functions for:

- Reading and merging enrollment and contact hour data
- Calculating original and generated FTE (Full-Time Equivalent) values
- Filtering and formatting reports by division, course, or instructor
- Computing enrollment percentages and course totals
- Exporting cleaned and structured data for reporting

Dependencies
------------
- pandas
- options4 (utility module)

Typical Use
-----------
This module is imported into the Streamlit `app.py` frontend as `wf`.

Example:
    import web_functions as wf
    df = wf.readfile()
    fte_df = wf.fte_by_div_raw(df, tier_df, 'ENG')
"""

import io
import pandas as pd
import options4 as opfour
from options4 import remove_duplicate_sections, calculate_enrollment_percentage, generate_fte
import re
# import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill

# import openpyxl
from openpyxl.utils import get_column_letter


# def readfile():
#     """
#     Reads, merges, and processes course and FTE data from CSV and Excel sources

#     This function:
#     - Reads 'deanDailyCsar.csv' and 'unique_deansDailyCsar_FTE.xlsx'
#     - Extracts 'Course Code' if missing
#     - Merges contact hours into the main dataset
#     - Computes total FTE per section

#     Returns
#     -------
#     pd.DataFrame
#         Cleaned and sorted merged dataset ready for FTE analysis,
#         or an empty list if files are missing.
#     """

#     try:
#         # reads the deansDailyCsar.csv and unique_deansDailyCsar_FTE files in to a dataframe
#         file_in = pd.read_csv('deanDailyCsar.csv')
#         fte_file_in = pd.read_excel('unique_deansDailyCsar_FTE.xlsx')

#         # merge prior dataframes
#         # Extract Course Code from Sec Name if not already done
#         if "Course Code" not in file_in.columns:
#             file_in["Course Code"] = file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

#         # Also create Course Code in credits_df
#         if "Course Code" not in fte_file_in.columns:
#             fte_file_in["Course Code"] = fte_file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

#         # Merge only needed columns from credits_df
#         merged_df = pd.merge(
#             file_in,
#             fte_file_in[["Course Code", "Contact Hours"]],
#             how='left',
#             on='Course Code'
#         )

#         merged_df["Contact Hours"] = pd.to_numeric(merged_df["Contact Hours"], errors='coerce')
#         merged_df["FTE Count"] = pd.to_numeric(merged_df["FTE Count"], errors='coerce')

#         # Calculate Total FTE
#         merged_df["Total FTE"] = ((merged_df["Contact Hours"] * 16 *
#                                    merged_df["FTE Count"]) / 512).round(3)

#         # sorts the dataframe by sec divisions, sec name
#         # and sec faculty info and assigns it to groups
#         groups = merged_df.sort_values(["Sec Divisions", "Sec Name", "Sec Faculty Info"])

#         return groups

#     except FileNotFoundError:
#         groups = []
#         print("File Missing!")
#         return groups
def auto_format_excel(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
        assert isinstance(ws, Worksheet)

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Formatted {filename} with auto-adjusted column widths")

    except Exception as err:
        print(f"Error formatting Excel file: {str(err)}")


def process_sec_divisions(file_in):
    print("\nAvailable Sec Divisions:\n")
    sec_group = sorted(file_in["Sec Divisions"].dropna().unique())
    for i in range(0, len(sec_group), 4):
        print(" ".join(f"-{x}" for x in sec_group[i:i + 4]))

    sec_input = input("\nEnter Sec Divisions separated by commas or 'ALL': ").upper().strip()
    divisions_to_process = sec_group if sec_input == "ALL" else [div.strip() for div in sec_input.split(",")]

    invalid_codes = [div for div in divisions_to_process if div not in sec_group]
    if invalid_codes:
        print(f"Invalid Sec Division(s): {', '.join(invalid_codes)}")
        return

    save_choice = input("Valid divisions found. Save report? (Y/N): ").strip().upper()
    if save_choice == 'Y':
        filename = save_report(file_in, divisions_to_process)
        auto_format_excel(filename)
    else:
        print("Report not saved.")

def save_report(df, divisions):
    filename = "sec_division_report.xlsx"
    filtered = df[df["Sec Divisions"].isin(divisions)]
    filtered.to_excel(filename, index=False)
    print(f"Report saved to {filename}")
    return filename

def generate_faculty_fte_report(dean_df, fte_tier, faculty_name):
    """
    Generates an FTE report for a specific faculty member.

    Parameters
    ----------
    dean_df : pd.DataFrame
        The cleaned and merged dean dataset with all sections.
    fte_tier : pd.DataFrame
        DataFrame mapping course prefixes to FTE sector multipliers.
    faculty_name : str
        Exact match of the faculty member's name from 'Sec Faculty Info'.

    Returns
    -------
    pd.DataFrame
        Formatted faculty FTE report including a summary row.
    float
        Total original FTE assigned to the faculty.
    float
        Total generated FTE for the faculty using multipliers.
    """

    # Filter for faculty
    faculty_df = dean_df[dean_df["Sec Faculty Info"] == faculty_name].copy()
    faculty_df = remove_duplicate_sections(faculty_df)

    # Select relevant columns
    cols_to_keep = [
        "Sec Name", "X Sec Delivery Method", "Meeting Times", "Capacity",
        "FTE Count", "Total FTE", "Sec Divisions"
    ]
    for col in cols_to_keep:
        if col not in faculty_df.columns:
            faculty_df[col] = ""

    faculty_df = faculty_df[cols_to_keep]

    # Add Enrollment Percentage
    faculty_df["Enrollment Per"] = calculate_enrollment_percentage(
        faculty_df["FTE Count"], faculty_df["Capacity"]
    )

    # Generate FTE
    faculty_df = generate_fte(faculty_df, fte_tier)

    # Add course code extraction
    faculty_df["Course Code"] = faculty_df["Sec Name"].str.extract(r"([A-Z]{3}-\d{3}[A-Z]?)")
    faculty_df["Instructor"] = faculty_name

    # Compute totals
    faculty_df["Total FTE"] = pd.to_numeric(faculty_df["Total FTE"], errors='coerce')
    faculty_df["Generated FTE"] = pd.to_numeric(faculty_df["Generated FTE"], errors='coerce')

    total_original = faculty_df["Total FTE"].sum()
    total_generated = faculty_df["Generated FTE"].sum()

    # Add summary row
    summary_row = pd.Series({
        "Instructor": "",
        "Course Code": "TOTAL",
        "Sec Name": "",
        "X Sec Delivery Method": "",
        "Meeting Times": "",
        "Capacity": "",
        "FTE Count": "",
        "Total FTE": total_original,
        "Sec Divisions": "",
        "Enrollment Per": "",
        "Generated FTE": total_generated
    })

    # Reorder columns for output
    final_cols = [
        "Instructor", "Course Code", "Sec Name", "X Sec Delivery Method",
        "Meeting Times", "Capacity", "FTE Count", "Total FTE", "Sec Divisions",
        "Enrollment Per", "Generated FTE"
    ]

    final_df = pd.concat([faculty_df[final_cols], pd.DataFrame([summary_row])],
                         ignore_index=True)

    return final_df, total_original, total_generated

def readfile(uploaded_file=None):
    """
    Reads, merges, and processes course and FTE data from CSV and Excel sources.
    If an uploaded_file is provided, it will be used instead of the default file.

    Parameters
    ----------
    uploaded_file : file-like object or None

    Returns
    -------
    pd.DataFrame
        Cleaned and sorted merged dataset ready for FTE analysis,
        or an empty DataFrame if file is missing or invalid.
    """

    try:
        if uploaded_file is not None:
            if uploaded_file.name.endswith('.csv'):
                file_in = pd.read_csv(uploaded_file)
            else:
                file_in = pd.read_excel(uploaded_file)
        else:
            file_in = pd.read_csv('deanDailyCsar.csv')

        fte_file_in = pd.read_excel('unique_deansDailyCsar_FTE.xlsx')

        if "Course Code" not in file_in.columns:
            file_in["Course Code"] = file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

        if "Course Code" not in fte_file_in.columns:
            fte_file_in["Course Code"] = fte_file_in["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")

        merged_df = pd.merge(
            file_in,
            fte_file_in[["Course Code", "Contact Hours"]],
            how='left',
            on='Course Code'
        )

        merged_df["Contact Hours"] = pd.to_numeric(merged_df["Contact Hours"], errors='coerce')
        merged_df["FTE Count"] = pd.to_numeric(merged_df["FTE Count"], errors='coerce')

        merged_df["Total FTE"] = ((merged_df["Contact Hours"] * 16 *
                                   merged_df["FTE Count"]) / 512).round(3)

        groups = merged_df.sort_values(["Sec Divisions", "Sec Name", "Sec Faculty Info"])
        return groups

    except Exception as e:
        print(f"Error loading file: {e}")
        return pd.DataFrame()

def calc_enrollment(row):
    """
    Calculates the enrollment percentage for a course section.

    Parameters
    ----------
    row : pd.Series
        A row from the DataFrame with 'Capacity' and 'FTE Count' fields.

    Returns
    -------
    str
        The enrollment percentage formatted as a string (e.g., "85.71%"),
        or "N/A%" if calculation is not possible.
    """

    try:
        cap = float(row["Capacity"])
        fte = float(row["FTE Count"])

        if cap == 0:
            return "0%"

        percentage = (fte / cap) * 100
        return f"{percentage:.2f}%"

    except (ValueError, TypeError, ZeroDivisionError):
        return "N/A%"


def fte_by_div_raw(file_in, fte_tier, div_code):
    """
    Computes raw and generated FTE totals for a given division.

    Parameters
    ----------
    file_in : pd.DataFrame
        Main dataset containing section-level details.
    fte_tier : pd.DataFrame
        FTE tier multipliers by prefix/course ID.
    div_code : str
        The division code to filter by (e.g., "ENG").

    Returns
    -------
    pd.DataFrame
        DataFrame of section-level and course-level FTE breakdowns.
    float
        Sum of original FTEs in the division.
    float
        Sum of generated FTEs using tier multipliers.
    """

    # Filter division
    div_code = div_code.upper()
    div_data = file_in[file_in['Sec Divisions'] == div_code].copy()

    if div_data.empty:
        return None, 0, 0

    # Create lookup for prefix/course ID → New Sector multiplier
    fte_lookup = {
        row['Prefix/Course ID']: row['New Sector']
        for _, row in fte_tier.iterrows()
        if pd.notna(row['Prefix/Course ID'])
    }

    # Extract course codes from Sec Name
    div_data['Course Code'] = div_data['Sec Name'].str.extract(r'([A-Z]+-\d+)')
    div_data = div_data.sort_values(['Course Code', 'Sec Name'])

    base_fte_value = 1926

    output_rows = []
    current_course = None
    course_total_fte = 0
    first_row = True

    grand_total_original_fte = 0
    grand_total_generated_fte = 0

    for _, row in div_data.iterrows():
        course = row['Course Code']
        sec = row['Sec Name'][:3] if pd.notna(row['Sec Name']) else ""

        new_sector_value = fte_lookup.get(sec, 0)
        total_fte = float(row['Total FTE']) if pd.notna(row['Total FTE']) else 0
        adjusted_fte = total_fte * (new_sector_value + base_fte_value)

        grand_total_original_fte += total_fte

        enrollment_per = ''
        if pd.notna(row['Capacity']) and pd.notna(row['FTE Count']) and float(row['Capacity']) > 0:
            enrollment_per = round((float(row['FTE Count']) / float(row['Capacity'])) * 100, 2)

        if course != current_course and current_course is not None:
            output_rows.append({
                'Division': '',
                'Course Code': 'Total',
                'Sec Name': '',
                'X Sec Delivery Method': '',
                'Meeting Times': '',
                'Capacity': '',
                'FTE Count': '',
                'Sec Faculty Info': '',
                'Total FTE': '',
                'Enrollment Per': '',
                'Generated FTE': course_total_fte
            })
            grand_total_generated_fte += course_total_fte
            course_total_fte = 0

        output_rows.append({
            'Division': div_code if first_row else '',
            'Course Code': course if course != current_course else '',
            'Sec Name': row['Sec Name'],
            'X Sec Delivery Method': row['X Sec Delivery Method'],
            'Meeting Times': row['Meeting Times'],
            'Capacity': row['Capacity'],
            'FTE Count': row['FTE Count'],
            'Sec Faculty Info': row['Sec Faculty Info'],
            'Total FTE': total_fte,
            'Enrollment Per': f"{enrollment_per}%" if enrollment_per != '' else '',
            'Generated FTE': adjusted_fte
        })

        course_total_fte += adjusted_fte
        current_course = course
        first_row = False

    # Add final course total
    if current_course is not None:
        output_rows.append({
            'Division': '',
            'Course Code': 'Total',
            'Sec Name': '',
            'X Sec Delivery Method': '',
            'Meeting Times': '',
            'Capacity': '',
            'FTE Count': '',
            'Sec Faculty Info': '',
            'Total FTE': '',
            'Enrollment Per': '',
            'Generated FTE': course_total_fte
        })
        grand_total_generated_fte += course_total_fte

    output_df = pd.DataFrame(output_rows)
    return output_df, grand_total_original_fte, grand_total_generated_fte


def format_fte_output(raw_df, original_fte_total, generated_fte_total):
    """
    Formats the FTE output DataFrame for display, including currency formatting

    Parameters
    ----------
    raw_df : pd.DataFrame
        Unformatted FTE data generated by `fte_by_div_raw`.
    original_fte_total : float
        Sum of original (unadjusted) FTEs.
    generated_fte_total : float
        Sum of adjusted/generated FTEs.

    Returns
    -------
    pd.DataFrame
        Formatted DataFrame with human-readable FTE values and a
        division total row.
    """

    formatted_rows = []

    for _, row in raw_df.iterrows():
        formatted_row = row.copy()
        if isinstance(row['Generated FTE'], (float, int)):
            formatted_row['Generated FTE'] = "${:,.3f}".format(row['Generated FTE'])
        formatted_rows.append(formatted_row)

    df = pd.DataFrame(formatted_rows)
    df.loc[len(df.index)] = {
        'Division': '',
        'Course Code': 'DIVISION TOTAL',
        'Sec Name': '',
        'X Sec Delivery Method': '',
        'Meeting Times': '',
        'Capacity': '',
        'FTE Count': '',
        'Sec Faculty Info': '',
        'Total FTE': '',
        'Enrollment Per': '',
        'Generated FTE': "${:,.2f}".format(generated_fte_total)
    }

    return df


def calculate_fte_by_course(df, fte_tier, course_code, base_fte=1926):
    """
    Computes FTE statistics for a specific course across all its sections.

    Parameters
    ----------
    df : pd.DataFrame
        The merged dataset with section-level data.
    fte_tier : pd.DataFrame
        Tier multipliers for generating FTE values.
    course_code : str
        The course code (e.g., "ENG-111") to filter.
    base_fte : int, optional
        Base FTE value used in generation, default is 1926.

    Returns
    -------
    pd.DataFrame
        Formatted DataFrame containing FTE and enrollment details by section.
    float
        Total original FTE for the course.
    float
        Total generated FTE using sector multipliers.
    """

    course_code = course_code.upper()
    filtered = df[df['Course Code'] == course_code].copy()
    filtered = filtered.drop_duplicates(subset='Sec Name')

    if filtered.empty:
        return None, 0, 0

    # Load FTE lookup
    fte_lookup = {
        row['Prefix/Course ID']: row['New Sector']
        for _, row in fte_tier.iterrows()
        if pd.notna(row['Prefix/Course ID'])
    }

    output_rows = []
    total_original_fte = 0
    total_generated_fte = 0

    for _, row in filtered.iterrows():
        sec_prefix = row['Sec Name'][:3]
        new_sector = fte_lookup.get(sec_prefix, 0)
        total_fte = float(row['Total FTE']) if pd.notna(row['Total FTE']) else 0
        generated_fte = total_fte * (new_sector + base_fte)
        total_original_fte += total_fte
        total_generated_fte += generated_fte

        enrollment_per = ''
        if pd.notna(row['Capacity']) and pd.notna(row['FTE Count']) and float(row['Capacity']) > 0:
            enrollment_per = round((float(row['FTE Count']) / float(row['Capacity'])) * 100, 2)

        output_rows.append({
            'Sec Name': row['Sec Name'],
            'X Sec Delivery Method': row['X Sec Delivery Method'],
            'Sec Faculty Info': row['Sec Faculty Info'],
            'Meeting Times': row['Meeting Times'],
            'Capacity': row['Capacity'],
            'FTE Count': row['FTE Count'],
            'Total FTE': total_fte,
            'Enrollment Per': f"{enrollment_per}%" if enrollment_per else '',
            'Generated FTE': generated_fte
        })

    # Add summary row
    output_rows.append({
        'Sec Name': 'COURSE TOTAL',
        'X Sec Delivery Method': '',
        'Sec Faculty Info': '',
        'Meeting Times': '',
        'Capacity': '',
        'FTE Count': '',
        'Total FTE': total_original_fte,
        'Enrollment Per': '',
        'Generated FTE': total_generated_fte
    })

    df_out = pd.DataFrame(output_rows)
    df_out['Generated FTE'] = df_out['Generated FTE'].apply(lambda x: "${:,.2f}".format(x) if isinstance(x, (float, int)) else x)

    return df_out, total_original_fte, total_generated_fte


def save_faculty_excel(data, instructor_name):
    """
    Generates an Excel report for a faculty member with structured formatting.

    Parameters
    ----------
    data : pd.DataFrame
        Cleaned and formatted faculty data.
    instructor_name : str
        Name of the instructor.
    
    Returns
    -------
    BytesIO
        A byte stream of the Excel file ready for Streamlit download.
    """

    # Create in-memory output
    output = io.BytesIO()

    # Create Excel writer
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        data.to_excel(writer, sheet_name='Faculty Report', index=False, startrow=1)

        workbook = writer.book
        worksheet = writer.sheets['Faculty Report']

        # Define formats
        header_format = workbook.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
        money_format = workbook.add_format({"num_format": "$#,##0.00"})
        number_format = workbook.add_format({"num_format": "#,##0.0"})
        total_format = workbook.add_format({"bold": True, "bg_color": "#E0E0E0", "border": 1})

        # Set custom header row
        for col_num, value in enumerate(data.columns):
            worksheet.write(0, col_num, value, header_format)

        # Format numeric columns
        for col_num, col_name in enumerate(data.columns):
            if "FTE" in col_name and "Generated" in col_name:
                worksheet.set_column(col_num, col_num, 15, money_format)
            elif "FTE" in col_name:
                worksheet.set_column(col_num, col_num, 12, number_format)
            elif col_name in ["Instructor", "Course Code", "Sec Name"]:
                worksheet.set_column(col_num, col_num, 20)
            else:
                worksheet.set_column(col_num, col_num, 18)

        # Add totals row
        total_row = len(data) + 1
        worksheet.write(total_row, 0, "TOTAL", total_format)

        if "Total FTE" in data.columns:
            worksheet.write_formula(total_row, data.columns.get_loc("Total FTE"),
                                    f'=SUM(H2:H{total_row})', total_format)

        if "Generated FTE" in data.columns:
            col_index = data.columns.get_loc("Generated FTE")
            col_letter = chr(65 + col_index)
            worksheet.write_formula(total_row, col_index,
                                    f'=SUM({col_letter}2:{col_letter}{total_row})',
                                    total_format)

    output.seek(0)
    return output
