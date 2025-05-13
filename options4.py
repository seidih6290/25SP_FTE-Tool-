import pandas as pd
import re
import os
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side


def clean_name_for_search(name):
    """
    Standardize name format for searching.
    Removes periods and extra spaces.

    Parameters
    ----------
    name : str
        Name to clean

    Returns
    -------
    str
        Cleaned name for comparison
    """
    return name.replace(".", "").strip().lower()


def clean_instructor_name(name):
    """
    Clean instructor name for file naming.
    Handles different formats like "H Seidi", "H. Seidi", etc.

    Parameters
    ----------
    name : str
        Instructor name

    Returns
    -------
    str
        Cleaned name formatted for filename
    """
    # Split by comma first if it exists
    if "," in name:
        last_name, first_part = name.split(",", 1)
        # Clean up the last name and first initial
        last_name = last_name.strip().lower()
        # Get first character and remove any periods
        first_initial = first_part.strip().replace(".", "")[0].lower()
    else:
        # Handle space-separated names
        parts = name.split()
        last_name = parts[-1].lower()
        # Get first character and remove any periods
        first_initial = parts[0].replace(".", "")[0].lower()

    return f"{last_name}{first_initial}_FTE.xlsx"


def fte_per_faculty(data):
    """Prompts user for faculty name and then creates an excel sheet
    with FTE information for the courses for that faculty member

    Parameters
    ----------
    faculty_data: pd.DataFrame
        DataFrame to extract information from
    course_tier: pd.DataFrame
        DataFrame to extract information from
    """

    print()
    # Get unique faculty names for reference
    faculty = sorted(data["Sec Faculty Info"].dropna().unique())
    print(f"Found {len(faculty)} faculty members")
    selected_faculty = None

    go = True
    while go:
        print("\nEnter instructor name (first or last name)")
        print("Type 'list' to see all instructors")
        print("Type 'back' for main menu")

        faculty_name = input("\nEnter name: ").strip()

        if faculty_name.lower() == "back":
            print("\nReturning to main menu...")
            return

        if faculty_name.lower() == "list":
            print("\nInstructors:")
            for i in range(0, len(faculty), 3):
                names = faculty[i: i + 3]
                print("  ".join(f"{name:<30}" for name in names))
            continue

        if not faculty_name:
            print("Please enter a valid name.")
            continue

        # Case-insensitive search for partial matches with standardized format
        cleaned_input = clean_name_for_search(faculty_name)
        matches = [f for f in faculty
                   if cleaned_input in clean_name_for_search(f)]
        print(f"Found {len(matches)} matching instructors")

        if not matches:
            print(f"No instructors found matching '{faculty_name}'.")
            print("Try searching without periods (.) or check the instructor"
                  " list.")

        if len(matches) > 1:
            print("\nMultiple instructors found:")
            for i, name in enumerate(matches, 1):
                print(f"{i}. {name}")

            choice = input("\nEnter number to select instructor "
                           " (or press Enter to search again): ")
            if not choice.isdigit() or int(choice) not in range(1, len(matches)):
                continue
            selected_faculty = matches[int(choice) - 1]
        else:
            selected_faculty = matches[0]

        go = False

    assert selected_faculty is not None
    print(f"\nProcessing data for: {selected_faculty}")

    name = selected_faculty

    # Read FTE tier data
    course_tier = pd.read_excel("FTE_Tier.xlsx")

    # Filter data for the selected faculty
    frame = data[data["Sec Faculty Info"] == selected_faculty].copy()

    frame = remove_duplicate_sections(frame)

    # Select columns of interest
    frame = frame[
        [
            "Sec Name",
            "X Sec Delivery Method",
            "Meeting Times",
            "Capacity",
            "FTE Count",
            "Total FTE",
            "Sec Divisions",
        ]
    ]

    # Get course codes
    courses = get_column_uniques(frame, "Sec Name")
    course_codes = sorted(get_course_codes(courses))

    # Create a safe filename
    name_parts = selected_faculty.split()
    if len(name_parts) >= 2:
        filename = name_parts[1] + name_parts[0][0]  # lastname + first initial
    else:
        filename = name_parts[0]  # fallback if only one name

    # Calculate enrollment percentage
    frame["Enrollment Per"] = calculate_enrollment_percentage(
        frame["FTE Count"], frame["Capacity"]
    )

    # Generate adjusted FTE if course tier info is available
    if course_tier is not None:
        frame = generate_fte(frame, course_tier)

    # Create the Excel file
    create_instructor_excel(
        data=frame,
        name=filename,
        course_codes=course_codes,
        instructor_name=selected_faculty,
    )


def create_instructor_excel(data, name, course_codes, instructor_name):
    """Creates an Excel report for instructor FTE data

    Parameters
    ----------
    data: pd.DataFrame
        Instructor data frame
    name: str
        Name for the Excel file
    course_codes: list
        List of course codes to include
    instructor_name: str
        Name of the instructor for the report
    """

    # Calculate totals

    # Create the Excel filename
    course_name = re.match(r"[A-Z]{3}-\d{3}[A-Z]?", name)
    if course_name is not None:
        filename = name.split("-")[0].lower() + name.split("-")[1]
    else:
        filename = name.lower()
    filename += "_FTE.xlsx"
    file_path = os.path.join(os.getcwd(), filename)

    # Create Excel workbook
    excel_options = {"nan_inf_to_errors": True}
    with xlsxwriter.Workbook(file_path, excel_options) as workbook:
        worksheet = workbook.add_worksheet()

        # Define formats
        header_format = workbook.add_format({"bold": True})
        number_format = workbook.add_format({"num_format": "#,##0.00"})
        money_format = workbook.add_format({"num_format": "$#,##0.00"})
        total_format = workbook.add_format({"bold": True,
                                            "bg_color": "#E0E0E0"})
        total_number_format = workbook.add_format(
            {"bold": True, "bg_color": "#E0E0E0", "num_format": "#,##0.00"}
        )
        total_money_format = workbook.add_format(
            {"bold": True, "bg_color": "#E0E0E0", "num_format": "$#,##0.00"}
        )

        # Write header row
        headers = [
            "Instructor",
            "Course Code",
            "Sec Name",
            "X Sec Delivery Method",
            "Meeting Times",
            "Capacity",
            "FTE Count",
            "Total FTE",
            "Sec Divisions",
        ]

        # Add Generated FTE column to headers
        if "Generated FTE" in data.columns:
            headers.append("Generated FTE")

        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Track current row
        current_row = 1

        # Write instructor name in first row, first column
        worksheet.write(current_row, 0, instructor_name)

        # Process each course
        grand_total_fte = 0
        grand_total_generated_fte = 0

        for course in course_codes:
            # Filter data for this course
            course_data = data[
                data["Sec Name"].str.contains(course, case=False, na=False)
            ].copy()
            course_data = course_data.sort_values(by=["Sec Name"])

            # Write course code in first column
            worksheet.write(current_row, 1, course)

            # Track course totals
            course_total_fte = 0
            course_total_generated_fte = 0

            # Process each section in the course
            for _, row in course_data.iterrows():
                # Calculate running totals
                course_total_fte += row["Total FTE"]
                if "Generated FTE" in data.columns:
                    course_total_generated_fte += row["Generated FTE"]

                # Write row data
                worksheet.write(current_row, 2, row["Sec Name"])
                worksheet.write(current_row, 3, row["X Sec Delivery Method"])
                worksheet.write(current_row, 4, row["Meeting Times"])
                worksheet.write(current_row, 5, row["Capacity"])
                worksheet.write(current_row, 6, row["FTE Count"])
                worksheet.write_number(current_row, 7, row["Total FTE"],
                                       number_format)

                # Handle #NUM! or missing values in Sec Divisions
                if (
                    pd.notna(row["Sec Divisions"])
                    and str(row["Sec Divisions"]) != "#NUM!"
                ):
                    worksheet.write(current_row, 8, row["Sec Divisions"])
                else:
                    worksheet.write(
                        current_row, 8, ""
                    )  # Write empty string for missing/error values

                # Write Generated FTE if available
                if "Generated FTE" in data.columns:
                    worksheet.write_number(
                        current_row, 9, row["Generated FTE"], money_format
                    )

                current_row += 1

            # Update grand totals
            grand_total_fte += course_total_fte
            grand_total_generated_fte += course_total_generated_fte

            # Write course total row
            worksheet.write(
                current_row, 1, "Total"
            )  # Put Total under Course Code column
            worksheet.write_number(current_row, 7, course_total_fte,
                                   number_format)

            # Write Generated FTE course total if available
            if "Generated FTE" in data.columns:
                worksheet.write_number(
                    current_row, 9, course_total_generated_fte, money_format
                )

            current_row += 1

        # Write grand total row
        for col in range(len(headers)):
            # Apply formatting to all cells in the Total row
            # Put Total in Course Code column (column B)
            if col == 1:
                worksheet.write(current_row, col, "Total", total_format)
            # Total FTE column
            elif col == 7:
                worksheet.write_number(
                    current_row, col, grand_total_fte, total_number_format
                )
            # Generated FTE column
            elif col == 9 and "Generated FTE" in data.columns:
                worksheet.write_number(
                    current_row, col, grand_total_generated_fte,
                    total_money_format
                )
            else:
                worksheet.write(current_row, col, "", total_format)

        # Format columns to appropriate width
        column_widths = {
            0: 15,  # Instructor
            1: 12,  # Course Code
            2: 15,  # Sec Name
            3: 20,  # X Sec Delivery Method
            4: 40,  # Meeting Times
            5: 10,  # Capacity
            6: 10,  # FTE Count
            7: 10,  # Total FTE
            8: 15,  # Sec Divisions
            9: 15,  # Generated FTE
        }

        for col, width in column_widths.items():
            if col < len(headers):
                worksheet.set_column(col, col, width)

    # Open with openpyxl to add bottom border above total
    wb = load_workbook(file_path)
    ws = wb.active

    assert isinstance(ws, Worksheet)
    # Get last row (where the total is)
    last_row = ws.max_row

    # Add a bottom border to cells before the total
    if last_row > 2:  # Make sure there are rows before the total
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=last_row - 1, column=col)
            cell.border = Border(bottom=Side(style="thin"))

    wb.save(file_path)

    # Print summary
    print(f"\nCreated file: {file_path}")
    print(f"Total FTE for {instructor_name}: {grand_total_fte:.2f}")
    if "Generated FTE" in data.columns:
        print(f"Total Generated FTE for {instructor_name}: "
              f"${grand_total_generated_fte:.2f}")


def get_course_frame(data, name, apply_filter=True):
    """Extracts rows associated with a course code

    Parameters
    ----------
    data: pd.DataFrame
        DataFrame to extract rows from.
    name: str
        Course code to filter for.
    filter: bool (default = True)
        If true, face-to-face courses will be filtered

    Returns
    -------
    pd.DataFrame
        All rows associated to the Course Code without face-to-face
        classes with INET meeting times if filtered, else all rows.
    """

    assert isinstance(data, pd.DataFrame)
    assert name is not None

    # Filter for matching course code
    frame = data[data["Sec Name"].str.contains(name, case=False,
                                               na=False)].copy()

    if apply_filter:
        frame = remove_duplicate_sections(frame)
    else:
        frame = remove_duplicate_sections(frame)

    return frame


def remove_duplicate_sections(frame):
    """
    Removes duplicate course section rows based on 'Sec Name'.
    Keeps the first entry for each section (regardless of delivery method).

    Parameters
    ----------
    frame : pd.DataFrame
        DataFrame with potential duplicate course sections.

    Returns
    -------
    pd.DataFrame
        Cleaned DataFrame with only one row per section.
    """

    assert isinstance(frame, pd.DataFrame)
    frame = frame.sort_values(by=["Sec Name", "Meeting Times"],
                              na_position="last")
    return frame.drop_duplicates(subset=["Sec Name"], keep="first")


def sort_dataframe(data, sort_by=["Sec Divisions", "Sec Name",
                                  "Sec Faculty Info"]):
    """Sorts a DataFrame by columns, in ascending order.

    Parameters
    ----------
    data: pd.DataFrame
        DataFrame to sort.
    sort_by: list[str]
        (default = ["Sec Divisions", "Sec Name", "Sec Faculty Info"])
        Column name(s) to sort by.

    Returns
    -------
    pd.DataFrame
        Sorted DataFrame
    """
    assert isinstance(data, pd.DataFrame)
    return data.sort_values(by=sort_by)


def total_ftes(data):
    """
    calculates to total FTE for each course and for a division
    :param data: ps.DataFrame
        A DataFrame that has individual secs generated FTE
    :return:
    course_fte: dictionary
        courses and their total generated FTE
    final_fte: Interger
        total generated FTE for entire dataframe
    """
    assert isinstance(data, pd.DataFrame)

    try:
        # Ensure required columns exist
        if "Sec Name" not in data.columns:
            raise KeyError("Missing required column: 'Sec Name'")
        if "Total FTE" not in data.columns:
            raise KeyError("Missing required column: 'Total FTE'")
        if "Generated FTE" not in data.columns:
            raise KeyError("Missing required column: 'Generated FTE'")

        # Ensure FTE column contains valid numeric values
        if not pd.api.types.is_numeric_dtype(data["Total FTE"]):
            raise ValueError("Column 'FTE' must contain only numeric values.")

        # Get the totals for different courses
        data["_Course Code"] = data["Sec Name"].str.extract(r"([A-Z]{3}-\d{3})")
        course_fte_total = (
            data.groupby("_Course Code")[("Generated FTE")].sum().to_dict()
        )
        data.drop(columns=["_Course Code"], inplace=True, errors="ignore")

        # Get total for the entire division
        final_fte_total = data["Generated FTE"].sum()

        return course_fte_total, final_fte_total

    except KeyError as e:
        print(f"️ KeyError in total_FTEs: {e}")
    except ValueError as e:
        print(f"️ ValueError in total_FTEs: {e}")
    except TypeError as e:
        print(f"️ TypeError in total_FTEs: {e}")

    return {}, 0


def generate_fte(data, tier, support=1926):
    """
    calculates generated FTE for a set of data and returns new dataframe
    containing generated fte

    Parameters
    ----------
    data: pd.DataFrame
        DataFrame to calculate generated FTE for

    tier: pd.DataFrame
        DataFrame that holds the proposed funding lever for different
        tiers

    Returns
    -------
    pd.DataFrame
        generate_fte: a new DataFrame that has the generated FTE
    """
    assert isinstance(data, pd.DataFrame)
    assert isinstance(tier, pd.DataFrame)

    # Constant value used for calculating FTE
    try:
        # Ensure valid dataframes
        if not isinstance(data, pd.DataFrame):
            raise TypeError("Parameter 'data' must be a pandas DataFrame.")
        if not isinstance(tier, pd.DataFrame):
            raise TypeError("Parameter 'tier' must be a pandas DataFrame.")

        # Check if required columns exist in 'tier'
        required_tier_columns = ["Prefix/Course ID", "New Sector"]
        for col in required_tier_columns:
            if col not in tier.columns:
                raise KeyError(f"Missing required column '{col}'"
                               " in tier DataFrame.")

        # Check if required columns exist in 'data'
        required_data_columns = ["Sec Name", "Total FTE"]
        for col in required_data_columns:
            if col not in data.columns:
                raise KeyError(f"Missing required column '{col}' in data"
                               " DataFrame.")

        # create a dictionary to hold the course ID and their proposed
        # funding
        data["_Course Prefix"] = data["Sec Name"].str[:3]
        data.loc[:, "_Course Prefix"] = data["_Course Prefix"].fillna("UNKNOWN")
        courseid_to_funding = {
            row["Prefix/Course ID"]: row["New Sector"] for
            _, row in tier.iterrows()
        }

        # Apply computed generated FTE to for all rows in original
        # DataFrame
        data["Generated FTE"] = data.apply(
            lambda row: compute_fte(row, courseid_to_funding, support), axis=1
        )
        data.drop(columns=["_Course Prefix"], inplace=True, errors="ignore")

        return data

    except TypeError as e:
        print(f"TypeError in generate_fte: {e}")
    except KeyError as e:
        print(f"KeyError in generate_fte: {e}")

    return data.copy()


def compute_fte(row, courseid_to_funding, support=1926):
    """
    Computes the generate FTE for a single row in a dataframe
    :param row: pd.Series
        a row from the data DataFrame
    :param courseid_to_funding: dict
        a dictionary for the course prefixes and their funding levels
    :param support: int, optional
        a fixed amount for institutional and academic support(default
        is 1926)
    :return: float
        the computed generated FTE value for the row
    """

    try:
        # Ensure required columns are in DataFrame
        if "Sec Name" not in row:
            raise KeyError("Missing required column: 'Sec Name'")
        if "Total FTE" not in row:
            raise KeyError("Missing required column: 'Total FTE'")

        # Ensure 'Course Code' is a string and has at least 3 characters
        course_code = row["Sec Name"]
        if not isinstance(course_code, str) or len(course_code) < 3:
            raise ValueError(f"Invalid course code: {course_code}")

        # Extract prefix (first 3 chars)
        course_prefix = row["Sec Name"][:3]

        # Ensure 'Total FTE' is a number
        total_fte = row["Total FTE"]
        if not isinstance(total_fte, (int, float)) or pd.isna(total_fte):
            raise ValueError(f"Invalid 'Total FTE' value: {total_fte}")

        # Get funding level and calculate generated FTE
        prop_fund = courseid_to_funding.get(course_prefix, 0)
        return (prop_fund + support) * total_fte

    except KeyError as e:
        print(f" KeyError in compute_fte: {e}")
    except ValueError as e:
        print(f" ValueError in compute_fte: {e}")
    except TypeError as e:
        print(f" TypeError in compute_fte: {e}")

    return 0  # Return 0 if an error occurs so program doesn't crash


def calculate_enrollment_percentage(count, capacity):
    """Calculates enrollment percentage based on course count and
    capacity

    Paramters
    ---------
    count: int
        Student enrolled in the course
    capacity: int
        Max number of students that can be enrolled
    """
    if isinstance(capacity, pd.Series):
        # Replace any 0 values in the Series with NaN to prevent
        # division errors
        capacity = capacity.replace(0, pd.NA)

    return ((count / capacity) * 100).round(1).astype(str) + "%"


def fte_faculty_submenu(faculty):
    """Prompt the user for input, validate, and return it.

    Parameters
    ----------
    faculty: list[str]
        List of valid faculty names.

    Returns
    -------
    str or None
        The faculty member name, or "To be Announced",
         or None if the user quits
    """
    keep_going = True
    while keep_going:
        # Prompt  user for faculty name to search for
        print(f"{'FTE by Faculty':-^2}\n")
        choice = (
            input(
                "Enter the first and/or last name of a faculty "
                "member,\nTBA for classes with no announced "
                "faculty,\n(or enter Q to quit) \n "
                ">>> "
            )
            .strip()
            .title()
        )
        if choice == "Q":
            return None
        if choice == "Tba":
            return "To be Announced"

        # call function to search for faculty
        found_name = find_faculty(choice, faculty)

        # If multiple matches are found, prompt user to choose
        if isinstance(found_name, list) and len(found_name) > 1:
            found_name.append("None of these")  # Add a "None" option

            print_menu("Did you mean", found_name)
            name_choice = get_menu_choice(len(found_name))

            # User selects "None of these"
            if name_choice == len(found_name) - 1:
                found_name = None
            else:
                found_name = found_name[name_choice]

        # If exactly one match, return it
        elif isinstance(found_name, list) and len(found_name) == 1:
            found_name = found_name[0]

        # If no matches, notify user and loop again
        if not found_name:
            print(f"{choice} could not be found, please check spelling\n")
            continue  # Restart loop

        return found_name  # Return found faculty name


def find_faculty(search_for, to_search):
    """Searches faculty names for a match and returns None or the name
    in a list

    Parameters
    ----------
    search_for: str
        The name to search for.
    to_search: list[str]
        The names to search.

    Returns
    -------
    None | list[str]
        If not found returns None. Else, returns a list of all matches.
    """

    # Search for an exact faculty match
    name = [n for n in to_search if search_for == n]

    # If no exact match, attempt to match last names
    if not name:
        try:
            search_last = search_for.split()[1]  # Extract last name
            name = [n for n in to_search if n.split()[1] == search_last]
        except IndexError:
            # If user only provided one name, check against last names
            # in the list
            name = [n for n in to_search if search_for == n.split()[1]]

    # If still no match, compare first initials
    if not name:
        name = [n for n in to_search if search_for[0] == n[0]]

    # Return None if no match found
    return name if name else None


def get_course_codes(courses):
    """Cuts the section portion out of course codes.

    Parameters
    ----------
    courses: list[str]
        List of courses with sections.

    Returns
    -------
    set[str]
        Set of course codes without sections.
    """
    course_code_pattern = r"^([A-Z]{3}-\d{3}[A-Z]?)"
    course_codes = set()
    for code in courses:
        course_code = re.match(course_code_pattern, code)
        if course_code is not None:
            course_codes.add(course_code[1])

    return course_codes


def get_menu_choice(amount_options):
    """Prompts user for choice, validates, returns choice.

    print_menu is 0 based indexing on the back side but presented in the more
    readable format for the end user by adding 1 to the option number as it is
    printed to the screen. (Option 0 is presented as Option 1)

    Parameters
    ----------
    amount_options: int
        The amount of options presented to the end user.

    Returns
    -------
    int
        User input - 1
    """
    user_input = 0
    while user_input not in range(1, amount_options + 1):
        try:
            user_input = int(
                input(f"Enter a number between 1 and {amount_options}: "))
        except ValueError:
            # Ignoring anything other than a number in the correct range
            ...
    # Need to remove the 1 added to the options when printed for
    # menus that aren't main
    return user_input - 1


def print_menu(header, options):
    """Print formatted menu

    Parameters
    ----------
    header: str
        menu header
    options: list[str]
        menu options
    """
    # Get the amount of digits in the length of  options (e.g. 15 = 2
    # digits).
    length_digits = len(str(len(options)))
    # Add 2 to account for the '. ' in the print loop.
    length_formatting = length_digits + 2
    # Get the length of the longest option for formatting.
    max_length = max(len(op) + length_formatting for op in options)
    if max_length < len(header):
        max_length = len(header) + length_digits

    print(f"{header:-^{max_length}}")
    for i, option in enumerate(options):
        print(f"{i + 1:>{length_digits}}. {option}")


def get_column_uniques(data, name):
    """Extracts unique values from a column within a DataFrame.

    Parameters
    ----------
    data: pd.DataFrame
        DataFrame to read from.
    name: str
        Name of the column to extract unique values.

    Returns
    -------
    list[str]
        List of unique values.
    """
    assert isinstance(data, pd.DataFrame)
    assert name is not None

    # Extract unique, non-null values
    unique_values = data[name].dropna().unique()
    # Explicit conversion to list[str] to prevent type errors
    return [str(x) for x in unique_values]
